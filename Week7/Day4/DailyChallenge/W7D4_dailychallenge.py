import os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill

os.chdir(r"C:\DI-Bootcamp\Week7\Day4\DailyChallenge")

# 1. Create a new Excel workbook and a sheet named "Grades"
wb = Workbook()
ws = wb.active
ws.title = "Grades"

# 2. Define the student data
data = {
    "Joe": {"math": 65, "science": 78, "english": 98, "gym": 89},
    "Bill": {"math": 55, "science": 72, "english": 87, "gym": 95},
    "Tim": {"math": 100, "science": 45, "english": 75, "gym": 92},
    "Sally": {"math": 30, "science": 25, "english": 45, "gym": 100},
    "Jane": {"math": 100, "science": 100, "english": 100, "gym": 60}
}

# 3. Add header row
headers = ["Student", "Math", "Science", "English", "Gym"]
ws.append(headers)

# Apply bold and colored formatting to the header row
header_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
for cell in ws[1]:
    cell.font = Font(bold=True)
    cell.fill = header_fill

# 4. Add student data rows
for student, grades in data.items():
    ws.append([
        student,
        grades["math"],
        grades["science"],
        grades["english"],
        grades["gym"]
    ])

# 5. Add average formulas to the last row
num_students = len(data)
avg_row = num_students + 2  # Leave one row of spacing

ws[f"A{avg_row}"] = "Averages"
ws[f"A{avg_row}"].font = Font(bold=True)

# Add formulas for each subject
for col in range(2, 6):  # Columns B to E
    col_letter = chr(64 + col)  # Convert 2 → B, 3 → C, etc.
    formula = f"=AVERAGE({col_letter}2:{col_letter}{num_students+1})"
    ws[f"{col_letter}{avg_row}"] = formula
    ws[f"{col_letter}{avg_row}"].font = Font(italic=True, color="006100")

# 6. Save the workbook
wb.save("gradebook.xlsx")
