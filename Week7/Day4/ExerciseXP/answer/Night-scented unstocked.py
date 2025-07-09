import sys
sys.path.append(r"C:\wiseowl\Python\Courseware\Lib\site-packages")

# work with Excel
from openpyxl import load_workbook

# create a new workbook
plant_book = load_workbook(filename=r"c:\\wiseowl\\python\\courseware\\files\\plants.xlsx")

# get reference to first (and only) sheet
plant_sheet = plant_book["Sheet1"]

# keep looping down cells in first column
this_plant = plant_sheet.cell(1,1)

int_check = 0
while True:

    # avoid infinite loops if bug
    int_check += 1
    if int_check > 100:
        print("Integer check triggered")
        break

    # go to next plant
    this_plant = this_plant.offset(1,0)

    # if no such plant, stop
    if this_plant.value == None:
        print("\nThe above plants are not in stock")
        break

    # if this plant is not in stock, say so
    stock_cell = this_plant.offset(0,7)
    if stock_cell.value == "No":
        print(this_plant.value)

# close down the file
plant_book.close()
