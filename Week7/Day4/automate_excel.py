from openpyxl import Workbook
from openpyxl.styles import Font

# 1. Создаем рабочую книгу и активный лист
wb = Workbook()
ws = wb.active

# 2. Данные для записи
data = [
    ['Product', 'Price'],
    ['Apple', 1.2],
    ['Banana', 0.8],
    ['Cherry', 2.5]
]

# 3. Запись данных в Excel с использованием цикла
for row in data:
    ws.append(row)

# 4. Применяем жирный шрифт к первой строке (заголовок)
for cell in ws[1]:  # ws[1] — это первая строка
    cell.font = Font(bold=True)

# 5. Сохраняем файл
wb.save('formatted_products.xlsx')
